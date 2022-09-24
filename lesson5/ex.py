from openpyxl import load_workbook, Workbook


book = load_workbook("python_22.xlsx")
page = book["ДЗ 2 мес"]
lst = page["C3:C13"]
marks = []

for row in lst:
    cell = row[0]
    value = cell.value
    if value is not None:
        mark = value
    else:
        mark = 0
    marks.append(mark)


sr_ar = sum(marks) / len(marks)
new_excel = Workbook()
print(sr_ar)
page1 = new_excel.active
page1["A1"] = "Средний балл по группе"
page1["B1"] = sr_ar
new_excel.save("new_excel.xlsx")




